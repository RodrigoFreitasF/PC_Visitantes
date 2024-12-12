from tkinter import Tk, Label, Entry, Button, font, messagebox, Frame
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import Reference, PieChart
from util.db import SQL
import os


class RelatorioExcelTela(Tk):
    def __init__(self, usuario_logado):
        super().__init__()

        self.usuario_logado = usuario_logado
        self.PADX = 20
        self.PADY = 10
        self.ROXO = "#662c92"
        self.iconbitmap("../ceub.ico")

        self.title("Relatório em Excel")
        self.resizable(False, False)

        self.font_title = font.Font(font="Helvetica 16 bold")
        self.font_label = font.Font(family="Helvetica", size=12, weight="bold")
        self.font_button = font.Font(family="Helvetica", size=12, weight="bold")

        self.lbl_title = Label(self, text="Relatório de Visitas", font=self.font_title, fg=self.ROXO)
        self.lbl_title.grid(row=0, column=0, columnspan=2, padx=self.PADX, pady=self.PADY, sticky="ew")

        self.frame_input = Frame(self)
        self.frame_input.grid(row=1, column=0, columnspan=2, pady=self.PADY, sticky="ew")

        self.label_ano = Label(self.frame_input, text="Ano:", font=self.font_label, fg=self.ROXO, anchor="w")
        self.label_ano.grid(row=0, column=0, padx=self.PADX, pady=self.PADY, sticky="w")

        self.entry_ano = Entry(self.frame_input, font="Helvetica 12")
        self.entry_ano.grid(row=0, column=1, padx=self.PADX, pady=self.PADY, sticky="ew")

        self.frame_input.grid_columnconfigure(0, weight=1)
        self.frame_input.grid_columnconfigure(1, weight=4)

        self.frame_buttons = Frame(self)
        self.frame_buttons.grid(row=2, column=0, columnspan=2, sticky="ew")

        self.btn_voltar = Button(
            self.frame_buttons,
            text="Voltar",
            font=self.font_button,
            bg=self.ROXO,
            fg="white",
            cursor="hand2",
            command=self.voltar
        )
        self.btn_voltar.grid(row=0, column=0, padx=self.PADX, pady=self.PADY, sticky="ew")

        self.btn_gerar = Button(
            self.frame_buttons,
            text="Gerar Relatório",
            font=self.font_button,
            bg=self.ROXO,
            fg="white",
            cursor="hand2",
            command=self.gerar_relatorio_excel
        )
        self.btn_gerar.grid(row=0, column=1, padx=self.PADX, pady=self.PADY, sticky="ew")

        self.frame_buttons.grid_columnconfigure(0, weight=1, uniform="buttons")
        self.frame_buttons.grid_columnconfigure(1, weight=1, uniform="buttons")

    def gerar_relatorio_excel(self):
        # obter o ano digitado
        ano = self.entry_ano.get()
        if not ano.isdigit():
            messagebox.showerror("Erro", "Por favor, insira um ano válido.")
            return

        ano = int(ano)

        sql = SQL(esquema='bd_gestao_visitantes')
        cmd_visitas_mes = f"""
                    SELECT MONTH(dta_visita) AS mes, COUNT(*) AS total_visitas
                    FROM ta_visitas
                    WHERE YEAR(dta_visita) = {ano}
                    GROUP BY MONTH(dta_visita)
                    ORDER BY mes;
                """

        visitas_mes = sql.get_list(cmd_visitas_mes)

        # dicionário de meses
        meses = {
            1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril', 5: 'Maio', 6: 'Junho',
            7: 'Julho', 8: 'Agosto', 9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
        }

        # definição da planilha
        wb = Workbook()
        ws = wb.active
        ws.title = f"Relatório {ano}"

        # configurar cabeçalho
        ws.append(["Mês", "Total de Visitas"])
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6A0DAD", end_color="6A0DAD", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # preencher os dados
        total_visitas_ano = 0

        for i in range(1, 13):
            total_visitas_mes = next((row['total_visitas'] for row in visitas_mes if row['mes'] == i), 0)
            ws.append([meses[i], total_visitas_mes])
            total_visitas_ano += total_visitas_mes

        # adicionar linha de total
        total_font = Font(bold=True)
        ws.append(["Total", total_visitas_ano])
        total_row = ws.max_row

        for cell in ws[total_row]:
            cell.font = total_font

        # ajustar largura das colunas
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # grafico
        chart = PieChart()
        chart.style = 6

        data = Reference(ws, min_col=2, min_row=1, max_row=13, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=13)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, "E2")

        arquivo_excel = f"relatorio_visitas_{ano}.xlsx"
        wb.save(arquivo_excel)

        messagebox.showinfo("Sucesso", f"Relatório salvo como '{arquivo_excel}'.")

        # abrir o arquivo automaticamente
        try:
            if os.name == 'nt':
                os.startfile(arquivo_excel)
            elif os.name == 'posix':
                os.system(f'open {arquivo_excel}')
        except Exception as e:
            print(f"Erro ao abrir o arquivo: {e}")
        self.destroy()

    def voltar(self):
        from menu_principal.tela_menu import MainMenu
        self.destroy()
        MainMenu(self.usuario_logado)


if __name__ == "__main__":
    app = RelatorioExcelTela()
    app.mainloop()
