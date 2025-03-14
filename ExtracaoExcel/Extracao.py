from openpyxl import Workbook
from openpyxl.chart import (
    PieChart,
    Reference
)
from util.db import SQL
from util import Converter as cv


class Visitantes:

    def __init__(self):
        self.arquivo = 'ExtracaoVisitantes.xlsx'
        self.idt_visitas = 0
        self.sql = SQL(esquema='bd_gestao_visitantes')

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Extração Anual de Visitantes"
        self.ws['A1'] = "Janeiro:"
        self.ws['A2'] = "Fevereiro:"
        self.ws['B2'] = "Março:"
        self.ws['C2'] = "Abril:"
        self.ws['D2'] = "Maio:"
        self.ws['E2'] = "Junho:"
        self.ws['F2'] = "Julho:"
        self.ws['G2'] = "Agosto:"
        self.ws['H2'] = "Setembro:"
        self.ws['I2'] = "Outubro:"
        self.ws['J2'] = "Novembro:"
        self.ws['K2'] = "Dezembro:"
        self.ws.column_dimensions['A'].width = 30
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 25
        self.ws.column_dimensions['F'].width = 15

    def buscar_visitantes(self, idt_visitas):
        cmd = "SELECT * FROM ta_visitas WHERE dta_visita BETWEEN %s AND %s;"
        idt_visitas = self.sql.get_date(cmd, [idt_visitas])
        return idt_visitas


        lista = self.sql.get_list(cmd, [dta_visita])
        return lista

    def gerar_planilha(self):
        # Buscar o Visitante
        Visitantes = self.buscar_visitantes(self.idt_visitas)
        self.ws['B1'] = Visitantes

        # Buscar a parte financeira
        total_geral = 0.0
        linha = 3
        lista = self.buscar_visitantes(self.idt_visitas)
        for obj in lista:
            self.ws.cell(row=linha, column=1, value=obj['idt_visitas'])
            self.ws.cell(row=linha, column=2, value=cv.mysql_to_bra(obj['dta_visita']))


            self.ws.cell(row=linha, column=7, value=obj['tot_col'])
            total_geral += float(obj['tot_col'])
            linha += 1

        self.ws.cell(row=linha, column=6, value="Total de Visitas")


        # Criar o gráfico de pizza
        pie_chart = PieChart()
        labels = Reference(self.ws, min_col=1, max_col=1, min_row=3, max_row=linha)
        data = Reference(self.ws, min_col=7, max_col=7, min_row=2, max_row=linha - 1)
        pie_chart.add_data(data, titles_from_data=True)
        pie_chart.set_categories(labels)

        # Adicionar o gráfico à planilha
        self.ws.add_chart(pie_chart, "B12")

    def salvar(self):
        self.wb.save(self.arquivo)


if __name__ == '__main__':
    Visitantes = Visitantes()
    Visitantes.gerar_planilha()
    Visitantes.salvar()